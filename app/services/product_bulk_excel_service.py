from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload, joinedload

from app.models import Category, Product, Seller, Variant, StockAudit
from app.services.audit import write_audit
from datetime import datetime

SHEET_NAME = "products_bulk_edit"
REFS_SHEET_NAME = "refs"

EXPORT_HEADERS = [
    "product_id",
    "variant_id",
    "product_name",
    "product_sku",
    "category_name",
    "seller_name",
    "unit",
    "product_is_active",
    "variant_name",
    "pack_size",
    "cost_price",
    "sale_price",
    "stock",
    "variant_is_active",
]

REQUIRED_HEADERS = EXPORT_HEADERS.copy()

MAX_EXCEL_ROWS = 50000

logger = logging.getLogger(__name__)


@dataclass
class BulkImportErrorItem:
    row: int
    message: str


@dataclass
class BulkImportWarningItem:
    row: int
    message: str


@dataclass
class BulkImportResult:
    rows_total: int = 0
    rows_success: int = 0
    rows_failed: int = 0
    rows_skipped_no_changes: int = 0

    products_updated: int = 0
    variants_updated: int = 0
    stock_updated: int = 0

    errors: list[BulkImportErrorItem] = field(default_factory=list)
    warnings: list[BulkImportWarningItem] = field(default_factory=list)

    def add_error(self, row: int, message: str) -> None:
        self.rows_failed += 1
        self.errors.append(BulkImportErrorItem(row=row, message=message))

    def add_warning(self, row: int, message: str) -> None:
        self.warnings.append(BulkImportWarningItem(row=row, message=message))

    def to_dict(self) -> dict[str, Any]:
        return {
            "rows_total": self.rows_total,
            "rows_success": self.rows_success,
            "rows_failed": self.rows_failed,
            "rows_skipped_no_changes": self.rows_skipped_no_changes,
            "products_updated": self.products_updated,
            "variants_updated": self.variants_updated,
            "stock_updated": self.stock_updated,
            "errors": [{"row": x.row, "message": x.message} for x in self.errors],
            "warnings": [{"row": x.row, "message": x.message} for x in self.warnings],
        }


@dataclass
class ParsedBulkRow:
    row_num: int

    product_id: int
    variant_id: int

    product_name: str
    product_sku: str | None
    category_name: str
    seller_name: str
    unit: str
    product_is_active: bool

    variant_name: str
    pack_size: int
    cost_price: Decimal
    sale_price: Decimal
    stock: int
    variant_is_active: bool


class ProductBulkExcelService:
    """
    Массовое редактирование существующего каталога через Excel.

    Основные принципы:
    - одна строка = один variant
    - обновляем только существующие product/variant
    - новые товары / варианты не создаем
    - category_name и seller_name допускаются только из существующих справочников
    - сначала полная валидация файла, потом изменения
    - один commit в самом конце
    """

    def __init__(self, db: Session):
        self.db = db

    # =========================================================
    # EXPORT
    # =========================================================
    def _today_str(self) -> str:
        return datetime.now().strftime("%Y_%m_%d")

    def export_catalog_excel(self, actor: dict[str, Any] | None = None) -> tuple[bytes, str]:
        """
        Выгружает каталог в Excel.

        Возвращает:
        - bytes файла
        - имя файла
        """
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

        refs_ws = wb.create_sheet(REFS_SHEET_NAME)

        ws.append(EXPORT_HEADERS)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        products = (
            self.db.query(Product)
            .options(
                joinedload(Product.category),
                joinedload(Product.seller),
                selectinload(Product.variants),
            )
            .order_by(Product.id.asc())
            .all()
        )

        for product in products:
            product_variants = sorted(product.variants, key=lambda v: v.id)

            for variant in product_variants:
                ws.append([
                    product.id,
                    variant.id,
                    product.name or "",
                    product.sku or "",
                    product.category.name if product.category else "",
                    product.seller.name if product.seller else "",
                    product.unit or "шт",
                    1 if bool(product.is_active) else 0,
                    variant.name or "",
                    variant.pack_size if variant.pack_size is not None else 1,
                    self._decimal_to_excel(variant.unit_price_net_cost),
                    self._decimal_to_excel(variant.unit_price),
                    variant.stock if variant.stock is not None else 0,
                    1 if (bool(product.is_active) and bool(variant.is_active)) else 0,
                ])

        self._build_refs_sheet(refs_ws)
        self._apply_sheet_style(ws)
        self._apply_validations(ws)

        # Скрытый лист со справочниками
        refs_ws.sheet_state = "hidden"

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"catalog_bulk_edit_{self._today_str()}.xlsx"
        return output.getvalue(), filename

    # =========================================================
    # IMPORT
    # =========================================================

    def import_catalog_excel(
        self,
        *,
        file_bytes: bytes,
        original_filename: str,
        actor: dict[str, Any] | None = None,
    ) -> dict[str, Any]:
        """
        Полный сценарий bulk import.

        Порядок:
        1. открыть workbook
        2. проверить лист и заголовки
        3. прочитать и провалидировать все строки
        4. проверить дубли variant_id в файле
        5. заранее загрузить products / variants / categories / sellers
        6. еще раз провалидировать связки
        7. применить изменения в памяти
        8. записать audit / stock_audit
        9. один commit в конце
        """
        actor = actor or {
            "user_id": None,
            "username": "system",
            "user_role": None,
        }

        result = BulkImportResult()

        logger.info("[BULK IMPORT] start filename=%s size_bytes=%s", original_filename, len(file_bytes))

        wb = load_workbook(filename=BytesIO(file_bytes), data_only=False)

        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Лист '{SHEET_NAME}' не найден в Excel-файле")

        ws = wb[SHEET_NAME]

        logger.info("[BULK IMPORT] sheet=%s max_row=%s max_column=%s", SHEET_NAME, ws.max_row, ws.max_column)

        if ws.max_row > MAX_EXCEL_ROWS:
            raise ValueError(f"Слишком много строк в Excel. Лимит: {MAX_EXCEL_ROWS}")

        header_map = self._read_headers(ws)
        self._validate_headers(header_map)

        logger.info("[BULK IMPORT] headers=%s", list(header_map.keys()))

        parsed_rows: list[ParsedBulkRow] = []
        seen_variant_ids: set[int] = set()

        # ---- 1. Первичный разбор и валидация строк ----
        for row_idx in range(2, ws.max_row + 1):
            raw = self._row_to_dict(ws, row_idx, header_map)

            if self._is_empty_row(raw):
                continue

            result.rows_total += 1

            row = self._parse_and_validate_row(raw, row_idx)

            if row.variant_id in seen_variant_ids:
                raise ValueError(
                    f"Дубликат variant_id={row.variant_id} внутри Excel-файла "
                    f"(строка {row.row_num})"
                )
            seen_variant_ids.add(row.variant_id)

            parsed_rows.append(row)

        if not parsed_rows:
            raise ValueError("В файле нет ни одной строки для обновления")

        logger.info(
            "[BULK IMPORT] parsed_rows=%s unique_products=%s unique_variants=%s",
            len(parsed_rows),
            len({r.product_id for r in parsed_rows}),
            len({r.variant_id for r in parsed_rows}),
        )

        # ---- 2. Грузим все нужное пачкой ----
        product_ids = {r.product_id for r in parsed_rows}
        variant_ids = {r.variant_id for r in parsed_rows}
        category_names = {r.category_name.casefold() for r in parsed_rows}
        seller_names = {r.seller_name.casefold() for r in parsed_rows}

        products = (
            self.db.query(Product)
            .filter(Product.id.in_(product_ids))
            .all()
        )
        variants = (
            self.db.query(Variant)
            .filter(Variant.id.in_(variant_ids))
            .all()
        )
        categories = (
            self.db.query(Category)
            .filter(func.lower(Category.name).in_(category_names))
            .all()
        )
        sellers = (
            self.db.query(Seller)
            .filter(func.lower(Seller.name).in_(seller_names))
            .all()
        )

        products_by_id = {p.id: p for p in products}
        variants_by_id = {v.id: v for v in variants}
        categories_by_name = {c.name.casefold(): c for c in categories}
        sellers_by_name = {s.name.casefold(): s for s in sellers}

        logger.info(
            "[BULK IMPORT] loaded products=%s variants=%s categories=%s sellers=%s",
            len(products_by_id),
            len(variants_by_id),
            len(categories_by_name),
            len(sellers_by_name),
        )

        # ---- 3. Строгая валидация связей до любых изменений ----
        row_errors: list[BulkImportErrorItem] = []

                # Проверка дубликатов SKU внутри файла после изменений
        # Пустые SKU не проверяем.
        future_sku_map: dict[str, int] = {}
        changed_sku_map: dict[str, int] = {}

        for row in parsed_rows:
            product = products_by_id.get(row.product_id)
            variant = variants_by_id.get(row.variant_id)
            category = categories_by_name.get(row.category_name.casefold())
            seller = sellers_by_name.get(row.seller_name.casefold())

            if not category:
                logger.warning(
                    "[BULK IMPORT DEBUG] category_not_found row=%s raw_category=%r normalized=%r available_sample=%s",
                    row.row_num,
                    row.category_name,
                    row.category_name.casefold(),
                    list(categories_by_name.keys())[:10],
                )

            if not seller:
                logger.warning(
                    "[BULK IMPORT DEBUG] seller_not_found row=%s raw_seller=%r normalized=%r available_sample=%s",
                    row.row_num,
                    row.seller_name,
                    row.seller_name.casefold(),
                    list(sellers_by_name.keys())[:10],
                )

            if not product:
                row_errors.append(BulkImportErrorItem(row=row.row_num, message=f"Товар product_id={row.product_id} не найден"))
                continue

            if not variant:
                row_errors.append(BulkImportErrorItem(row=row.row_num, message=f"Вариант variant_id={row.variant_id} не найден"))
                continue

            if variant.product_id != row.product_id:
                row_errors.append(
                    BulkImportErrorItem(
                        row=row.row_num,
                        message=(
                            f"variant_id={row.variant_id} не принадлежит "
                            f"product_id={row.product_id}"
                        ),
                    )
                )
                continue

            if not category:
                row_errors.append(
                    BulkImportErrorItem(
                        row=row.row_num,
                        message=f"Категория '{row.category_name}' не найдена",
                    )
                )

            if not seller:
                row_errors.append(
                    BulkImportErrorItem(
                        row=row.row_num,
                        message=f"Продавец '{row.seller_name}' не найден",
                    )
                )

            if row.product_sku:
                sku_key = row.product_sku.casefold()
                existing_product_id = future_sku_map.get(sku_key)
                if existing_product_id and existing_product_id != row.product_id:
                    row_errors.append(
                        BulkImportErrorItem(
                            row=row.row_num,
                            message=f"SKU '{row.product_sku}' повторяется для разных товаров внутри файла",
                        )
                    )
                else:
                    future_sku_map[sku_key] = row.product_id

                # ✅ SKU проверяем на конфликт с БД только если пользователь реально меняет SKU
                current_product_sku = (product.sku or "").casefold()
                if sku_key != current_product_sku:
                    changed_existing_product_id = changed_sku_map.get(sku_key)
                    if changed_existing_product_id and changed_existing_product_id != row.product_id:
                        row_errors.append(
                            BulkImportErrorItem(
                                row=row.row_num,
                                message=f"SKU '{row.product_sku}' повторяется для разных товаров среди измененных строк",
                            )
                        )
                    else:
                        changed_sku_map[sku_key] = row.product_id

        # Проверка конфликта SKU с БД только для реально измененных SKU
        if changed_sku_map:
            sku_conflicts = (
                self.db.query(Product)
                .filter(func.lower(Product.sku).in_(list(changed_sku_map.keys())))
                .all()
            )
            for db_product in sku_conflicts:
                if not db_product.sku:
                    continue

                sku_key = db_product.sku.casefold()
                incoming_product_id = changed_sku_map.get(sku_key)

                if incoming_product_id and incoming_product_id != db_product.id:
                    conflict_row = next(
                        (
                            r.row_num
                            for r in parsed_rows
                            if r.product_sku
                            and r.product_sku.casefold() == sku_key
                            and r.product_id == incoming_product_id
                        ),
                        0,
                    )
                    row_errors.append(
                        BulkImportErrorItem(
                            row=conflict_row,
                            message=f"SKU '{db_product.sku}' уже используется другим товаром в базе (product_id={db_product.id})",
                        )
                    )

        if row_errors:
            for err in row_errors:
                logger.error("[BULK IMPORT ERROR] row=%s message=%s", err.row, err.message)
                result.add_error(err.row, err.message)

            logger.error("[BULK IMPORT] validation_failed total_errors=%s filename=%s", len(row_errors), original_filename)
            raise ValueError("Файл содержит ошибки. Исправьте Excel и загрузите снова.")

        # ---- 4. Применение изменений ----
        changed_product_ids: set[int] = set()
        changed_variant_ids: set[int] = set()

        for row in parsed_rows:
            product = products_by_id[row.product_id]
            variant = variants_by_id[row.variant_id]
            category = categories_by_name[row.category_name.casefold()]
            seller = sellers_by_name[row.seller_name.casefold()]

            product_old, product_new = self._apply_product_changes(
                product=product,
                row=row,
                category=category,
                seller=seller,
            )

            variant_old, variant_new, stock_changed = self._apply_variant_changes(
                variant=variant,
                row=row,
                actor=actor,
            )

            changed_any = False

            if product_old:
                changed_any = True
                changed_product_ids.add(product.id)
                write_audit(
                    self.db,
                    entity_type="product",
                    entity_id=product.id,
                    action="product_bulk_updated",
                    actor=actor,
                    old_data=product_old,
                    new_data=product_new,
                    note="Массовое обновление из Excel",
                )

            if variant_old:
                changed_any = True
                changed_variant_ids.add(variant.id)
                write_audit(
                    self.db,
                    entity_type="variant",
                    entity_id=variant.id,
                    action="variant_bulk_updated",
                    actor=actor,
                    old_data=variant_old,
                    new_data=variant_new,
                    note="Массовое обновление из Excel",
                )

            if stock_changed:
                result.stock_updated += 1

            if changed_any:
                result.rows_success += 1
            else:
                result.rows_skipped_no_changes += 1

        # ---- 5. Общий batch audit ----
        write_audit(
            self.db,
            entity_type="product_bulk_edit",
            entity_id=None,
            action="bulk_import",
            actor=actor,
            old_data=None,
            new_data={
                "filename": original_filename,
                "rows_total": result.rows_total,
                "rows_success": result.rows_success,
                "rows_skipped_no_changes": result.rows_skipped_no_changes,
                "products_updated": len(changed_product_ids),
                "variants_updated": len(changed_variant_ids),
                "stock_updated": result.stock_updated,
            },
            note="Массовое обновление каталога через Excel",
        )

        logger.info(
            "[BULK IMPORT] ready_to_commit rows_total=%s rows_success=%s rows_skipped=%s products_updated=%s variants_updated=%s stock_updated=%s",
            result.rows_total,
            result.rows_success,
            result.rows_skipped_no_changes,
            len(changed_product_ids),
            len(changed_variant_ids),
            result.stock_updated,
        )

        self.db.commit()

        logger.info("[BULK IMPORT] commit_success filename=%s", original_filename)

        result.products_updated = len(changed_product_ids)
        result.variants_updated = len(changed_variant_ids)

        return result.to_dict()

    # =========================================================
    # APPLY CHANGES
    # =========================================================

    def _apply_product_changes(
        self,
        *,
        product: Product,
        row: ParsedBulkRow,
        category: Category,
        seller: Seller,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        old_data: dict[str, Any] = {}
        new_data: dict[str, Any] = {}

        def track(field_name: str, old_value: Any, new_value: Any) -> None:
            if old_value != new_value:
                old_data[field_name] = old_value
                new_data[field_name] = new_value

        track("name", product.name, row.product_name)
        track("sku", product.sku, row.product_sku)
        track("category_id", product.category_id, category.id)
        track("seller_id", product.seller_id, seller.id)
        track("unit", product.unit, row.unit)
        track("is_active", bool(product.is_active), row.product_is_active)

        if old_data:
            product.name = row.product_name
            product.sku = row.product_sku
            product.category_id = category.id
            product.seller_id = seller.id
            product.unit = row.unit
            product.is_active = row.product_is_active

        return old_data, new_data

    def _apply_variant_changes(
        self,
        *,
        variant: Variant,
        row: ParsedBulkRow,
        actor: dict[str, Any],
    ) -> tuple[dict[str, Any], dict[str, Any], bool]:
        old_data: dict[str, Any] = {}
        new_data: dict[str, Any] = {}
        stock_changed = False

        def track(field_name: str, old_value: Any, new_value: Any) -> None:
            if old_value != new_value:
                old_data[field_name] = old_value
                new_data[field_name] = new_value

        old_stock = int(variant.stock or 0)
        new_stock = int(row.stock)

        # 🔹 если товар выключен, вариант тоже всегда выключаем
        final_variant_is_active = row.variant_is_active if row.product_is_active else False

        track("name", variant.name, row.variant_name)
        track("pack_size", int(variant.pack_size or 1), row.pack_size)
        track("unit_price_net_cost", self._to_decimal_str(variant.unit_price_net_cost), self._to_decimal_str(row.cost_price))
        track("unit_price", self._to_decimal_str(variant.unit_price), self._to_decimal_str(row.sale_price))
        track("stock", old_stock, new_stock)
        track("is_active", bool(variant.is_active), final_variant_is_active)

        if old_data:
            variant.name = row.variant_name
            variant.pack_size = row.pack_size
            variant.unit_price_net_cost = row.cost_price
            variant.unit_price = row.sale_price
            variant.stock = new_stock
            variant.is_active = final_variant_is_active

        if old_stock != new_stock:
            stock_changed = True

            self.db.add(
                StockAudit(
                    variant_id=variant.id,
                    change_type="SET",
                    delta_units=new_stock - old_stock,
                    old_stock=old_stock,
                    new_stock=new_stock,
                    boxes=None,
                    units_per_box=None,
                    extra_units=None,
                    note="Массовое обновление из Excel",
                    user=(actor.get("username") or "admin")[:64],
                )
            )

        return old_data, new_data, stock_changed

    # =========================================================
    # EXCEL BUILDERS
    # =========================================================

    def _build_refs_sheet(self, ws) -> None:
        ws["A1"] = "categories"
        ws["B1"] = "sellers"
        ws["C1"] = "bool_values"

        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)

        categories = self.db.query(Category).order_by(Category.name.asc()).all()
        sellers = self.db.query(Seller).order_by(Seller.name.asc()).all()

        row = 2
        for category in categories:
            ws.cell(row=row, column=1, value=category.name)
            row += 1

        row = 2
        for seller in sellers:
            ws.cell(row=row, column=2, value=seller.name)
            row += 1

        ws["C2"] = 1
        ws["C3"] = 0

    def _apply_sheet_style(self, ws) -> None:
        widths = {
            "A": 12,
            "B": 12,
            "C": 36,
            "D": 20,
            "E": 24,
            "F": 24,
            "G": 12,
            "H": 16,
            "I": 24,
            "J": 12,
            "K": 14,
            "L": 14,
            "M": 12,
            "N": 16,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _apply_validations(self, ws) -> None:
        # Категории
        category_validation = DataValidation(
            type="list",
            formula1=f"={REFS_SHEET_NAME}!$A$2:$A$5000",
            allow_blank=False,
        )
        ws.add_data_validation(category_validation)
        category_validation.add(f"E2:E{max(ws.max_row, 500)}")

        # Продавцы
        seller_validation = DataValidation(
            type="list",
            formula1=f"={REFS_SHEET_NAME}!$B$2:$B$5000",
            allow_blank=False,
        )
        ws.add_data_validation(seller_validation)
        seller_validation.add(f"F2:F{max(ws.max_row, 500)}")

        # Bool
        bool_validation = DataValidation(
            type="list",
            formula1=f"={REFS_SHEET_NAME}!$C$2:$C$3",
            allow_blank=False,
        )
        ws.add_data_validation(bool_validation)
        bool_validation.add(f"H2:H{max(ws.max_row, 500)}")
        bool_validation.add(f"N2:N{max(ws.max_row, 500)}")

    # =========================================================
    # VALIDATION / PARSING
    # =========================================================

    def _read_headers(self, ws) -> dict[str, int]:
        header_map: dict[str, int] = {}

        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=1, column=col_idx).value
            header = str(value).strip() if value is not None else ""

            if not header:
                continue

            if header in header_map:
                raise ValueError(f"Дублирующийся заголовок колонки: '{header}'")

            header_map[header] = col_idx

        return header_map

    def _validate_headers(self, header_map: dict[str, int]) -> None:
        missing_headers = [h for h in REQUIRED_HEADERS if h not in header_map]
        if missing_headers:
            raise ValueError(
                "В Excel отсутствуют обязательные колонки: " + ", ".join(missing_headers)
            )

    def _row_to_dict(self, ws, row_idx: int, header_map: dict[str, int]) -> dict[str, Any]:
        return {
            header: ws.cell(row=row_idx, column=col_idx).value
            for header, col_idx in header_map.items()
        }

    def _is_empty_row(self, raw: dict[str, Any]) -> bool:
        for value in raw.values():
            if value is None:
                continue
            if isinstance(value, str) and not value.strip():
                continue
            return False
        return True

    def _parse_and_validate_row(self, raw: dict[str, Any], row_num: int) -> ParsedBulkRow:
        product_id = self._parse_required_int(raw.get("product_id"), "product_id", row_num)
        variant_id = self._parse_required_int(raw.get("variant_id"), "variant_id", row_num)

        product_name = self._parse_required_str(raw.get("product_name"), "product_name", row_num, max_len=255)
        product_sku = self._parse_optional_str(raw.get("product_sku"), "product_sku", row_num, max_len=64)
        category_name = self._parse_required_str(raw.get("category_name"), "category_name", row_num, max_len=120)
        seller_name = self._parse_required_str(raw.get("seller_name"), "seller_name", row_num, max_len=120)
        unit = self._parse_required_str(raw.get("unit"), "unit", row_num, max_len=32)
        product_is_active = self._parse_bool(raw.get("product_is_active"), "product_is_active", row_num)

        variant_name = self._parse_required_str(raw.get("variant_name"), "variant_name", row_num, max_len=120)
        pack_size = self._parse_required_int(raw.get("pack_size"), "pack_size", row_num)
        if pack_size < 1:
            raise ValueError(f"Строка {row_num}: pack_size должен быть >= 1")

        cost_price = self._parse_decimal(raw.get("cost_price"), "cost_price", row_num)
        sale_price = self._parse_decimal(raw.get("sale_price"), "sale_price", row_num)
        stock = self._parse_required_int(raw.get("stock"), "stock", row_num)
        variant_is_active = self._parse_bool(raw.get("variant_is_active"), "variant_is_active", row_num)

        if cost_price < Decimal("0"):
            raise ValueError(f"Строка {row_num}: cost_price не может быть отрицательной")
        if sale_price < Decimal("0"):
            raise ValueError(f"Строка {row_num}: sale_price не может быть отрицательной")
        if stock < 0:
            raise ValueError(f"Строка {row_num}: stock не может быть отрицательным")

        return ParsedBulkRow(
            row_num=row_num,
            product_id=product_id,
            variant_id=variant_id,
            product_name=product_name,
            product_sku=product_sku,
            category_name=category_name,
            seller_name=seller_name,
            unit=unit,
            product_is_active=product_is_active,
            variant_name=variant_name,
            pack_size=pack_size,
            cost_price=cost_price,
            sale_price=sale_price,
            stock=stock,
            variant_is_active=variant_is_active,
        )

    # =========================================================
    # PARSE HELPERS
    # =========================================================

    def _parse_required_str(
        self,
        value: Any,
        field_name: str,
        row_num: int,
        *,
        max_len: int,
    ) -> str:
        text = "" if value is None else str(value).strip()
        if not text:
            raise ValueError(f"Строка {row_num}: поле '{field_name}' обязательно")
        if len(text) > max_len:
            raise ValueError(f"Строка {row_num}: поле '{field_name}' длиннее {max_len} символов")
        return text

    def _parse_optional_str(
        self,
        value: Any,
        field_name: str,
        row_num: int,
        *,
        max_len: int,
    ) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        if len(text) > max_len:
            raise ValueError(f"Строка {row_num}: поле '{field_name}' длиннее {max_len} символов")
        return text

    def _parse_required_int(self, value: Any, field_name: str, row_num: int) -> int:
        if value is None or (isinstance(value, str) and not value.strip()):
            raise ValueError(f"Строка {row_num}: поле '{field_name}' обязательно")

        try:
            if isinstance(value, float) and value.is_integer():
                return int(value)
            return int(str(value).strip())
        except Exception:
            raise ValueError(f"Строка {row_num}: поле '{field_name}' должно быть целым числом")

    def _parse_decimal(self, value: Any, field_name: str, row_num: int) -> Decimal:
        if value is None or (isinstance(value, str) and not value.strip()):
            raise ValueError(f"Строка {row_num}: поле '{field_name}' обязательно")

        try:
            normalized = str(value).replace(" ", "").replace(",", ".").strip()
            return Decimal(normalized)
        except (InvalidOperation, ValueError):
            raise ValueError(f"Строка {row_num}: поле '{field_name}' должно быть числом")

    def _parse_bool(self, value: Any, field_name: str, row_num: int) -> bool:
        if isinstance(value, bool):
            return value

        if isinstance(value, int):
            if value == 1:
                return True
            if value == 0:
                return False

        if isinstance(value, float):
            if value == 1.0:
                return True
            if value == 0.0:
                return False

        normalized = "" if value is None else str(value).strip().casefold()

        true_values = {"1", "true", "yes", "y", "да", "активен"}
        false_values = {"0", "false", "no", "n", "нет", "неактивен"}

        if normalized in true_values:
            return True
        if normalized in false_values:
            return False

        raise ValueError(
            f"Строка {row_num}: поле '{field_name}' должно быть 1/0, true/false, да/нет"
        )

    # =========================================================
    # SMALL HELPERS
    # =========================================================

    def _decimal_to_excel(self, value: Any) -> float:
        if value is None:
            return 0.0
        return float(value)

    def _to_decimal_str(self, value: Any) -> str:
        if value is None:
            return "0.00"
        return str(Decimal(value).quantize(Decimal("0.01")))

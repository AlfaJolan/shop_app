from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from app.models import (
    Category,
    Product,
    Seller,
    Variant,
    ProductImage,
    ProductVideo,
)
from app.utils.media import (
    extract_youtube_id,
    download_image_bytes,
    save_main_image_bytes,
    save_gallery_image_bytes,
    delete_main_image_file,
    clear_product_gallery_dir,
)


SHEET_NAME = "products_import"

# Это не "обязательные к заполнению" поля,
# а обязательные колонки, которые должны существовать в шаблоне Excel.
REQUIRED_HEADERS = [
    "product_sku",
    "product_name",
    "product_description",
    "category_name",
    "seller_name",
    "seller_city",
    "unit",
    "main_image_url",
    "gallery_urls",
    "youtube_url",
    "variant_name",
    "pack_size",
    "cost_price",
    "sale_price",
    "stock",
    "is_active",
]


@dataclass
class ImportErrorItem:
    row: int
    message: str


@dataclass
class ImportResult:
    # ---- Общая статистика по строкам ----
    rows_total: int = 0
    rows_success: int = 0
    rows_failed: int = 0

    # ---- Статистика по сущностям каталога ----
    products_created: int = 0
    products_updated: int = 0
    variants_created: int = 0
    variants_updated: int = 0
    categories_created: int = 0
    sellers_created: int = 0

    # ---- Медиа-статистика ----
    media_skipped: int = 0
    media_processed: int = 0

    # ---- Ошибки и предупреждения ----
    # errors = критичные ошибки строки (строка не импортирована)
    # warnings = некритичные ошибки, например не скачалось фото
    errors: list[ImportErrorItem] = field(default_factory=list)
    warnings: list[ImportErrorItem] = field(default_factory=list)

    def add_error(self, row: int, message: str) -> None:
        self.rows_failed += 1
        self.errors.append(ImportErrorItem(row=row, message=message))

    def add_warning(self, row: int, message: str) -> None:
        self.warnings.append(ImportErrorItem(row=row, message=message))


@dataclass
class ParsedRow:
    # ---- Product ----
    product_sku: str
    product_name: str
    product_description: str | None
    category_name: str
    seller_name: str
    seller_city: str | None
    unit: str

    # ---- Media ----
    main_image_url: str | None
    gallery_urls: list[str]
    youtube_url: str | None

    # ---- Variant ----
    variant_name: str
    pack_size: int
    cost_price: Decimal
    sale_price: Decimal
    stock: int
    is_active: bool


class ProductImportService:
    """
    Сервис импорта товаров из Excel.

    Что делает:
    - читает Excel `products_import`
    - валидирует строки
    - создает/обновляет Category
    - создает/обновляет Seller
    - создает/обновляет Product по SKU
    - создает/обновляет Variant по (product_id, variant_name, pack_size)

    Что делает по медиа в этой версии:
    - скачивает главное фото товара
    - сохраняет галерею товара
    - сохраняет YouTube-видео
    - обрабатывает медиа только ОДИН РАЗ на один SKU в рамках одного файла

    Важная логика:
    - если строка невалидна → строка падает в errors
    - если товар/вариант импортировались, но медиа не скачалось →
      это идет в warnings, а не ломает весь импорт
    """

    def __init__(self, db: Session):
        self.db = db

        # Кеш категорий, продавцов и товаров нужен,
        # чтобы не делать лишние SELECT на одинаковых значениях внутри одного файла.
        self._category_cache: dict[str, Category] = {}
        self._seller_cache: dict[str, Seller] = {}
        self._product_cache: dict[str, Product] = {}

        # Набор SKU, для которых медиа уже обработано.
        # Это важно, потому что один и тот же товар может идти в Excel в нескольких строках
        # (по одной строке на каждый вариант), а картинки/видео у него общие.
        self._media_processed_skus: set[str] = set()

    # =========================================================
    # Публичные методы
    # =========================================================

    def import_from_upload(self, upload_file) -> ImportResult:
        """
        Точка входа для FastAPI UploadFile.
        Считываем байты файла и передаем в общий импорт из bytes.
        """
        content = upload_file.file.read()
        return self.import_from_bytes(content)

    def import_from_bytes(self, content: bytes) -> ImportResult:
        """
        Основной сценарий импорта:
        1. открыть workbook
        2. проверить наличие листа
        3. проверить заголовки
        4. пройти по строкам
        5. валидировать и импортировать сущности
        6. отдельно обработать медиа
        """
        result = ImportResult()
        wb = load_workbook(filename=BytesIO(content), data_only=True)

        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"Лист '{SHEET_NAME}' не найден в Excel-файле")

        ws = wb[SHEET_NAME]
        header_map = self._read_headers(ws)

        missing_headers = [h for h in REQUIRED_HEADERS if h not in header_map]
        if missing_headers:
            raise ValueError(
                "В Excel отсутствуют обязательные колонки: " + ", ".join(missing_headers)
            )

        # Контроль дублей вариантов внутри одного Excel-файла.
        # Один и тот же (SKU + variant_name + pack_size) не должен повторяться.
        seen_variant_keys: set[tuple[str, str, int]] = set()

        for row_idx in range(2, ws.max_row + 1):
            raw = self._row_to_dict(ws, row_idx, header_map)

            # Пропускаем полностью пустые строки
            if self._is_empty_row(raw):
                continue

            result.rows_total += 1

            try:
                # 1. Парсим и валидируем строку
                row = self._parse_and_validate_row(raw)

                # 2. Проверяем дубликат варианта внутри текущего Excel
                variant_key = (
                    row.product_sku.casefold(),
                    row.variant_name.casefold(),
                    row.pack_size,
                )
                if variant_key in seen_variant_keys:
                    raise ValueError(
                        "Дубликат варианта внутри файла: одинаковые product_sku + variant_name + pack_size"
                    )
                seen_variant_keys.add(variant_key)

                # 3. Создаем / получаем справочники
                category, category_created = self._get_or_create_category(row.category_name)
                seller, seller_created = self._get_or_create_seller(
                    row.seller_name,
                    row.seller_city,
                )

                # 4. Создаем / обновляем товар
                product, product_created = self._get_or_create_product(
                    row,
                    category.id,
                    seller.id,
                )

                # 5. Создаем / обновляем вариант
                _, variant_created = self._get_or_create_variant(product.id, row)

                # 6. Обрабатываем медиа только один раз на SKU за весь импорт
                self._process_product_media_once(product, row, row_idx, result)

                # 7. Обновляем статистику
                if category_created:
                    result.categories_created += 1
                if seller_created:
                    result.sellers_created += 1
                if product_created:
                    result.products_created += 1
                else:
                    result.products_updated += 1
                if variant_created:
                    result.variants_created += 1
                else:
                    result.variants_updated += 1

                # 8. Коммитим именно по строке.
                # Это удобно: если следующая строка упадет, предыдущие уже не потеряются.
                self.db.commit()
                result.rows_success += 1

            except Exception as exc:
                self.db.rollback()
                result.add_error(row_idx, str(exc))

        return result

    # =========================================================
    # Работа с Excel
    # =========================================================

    def _read_headers(self, ws) -> dict[str, int]:
        """
        Считывает первую строку Excel как заголовки
        и возвращает словарь:
            {normalized_header: column_index}
        """
        header_map: dict[str, int] = {}

        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            if cell_value is None:
                continue

            normalized = self._normalize_header(str(cell_value))
            if normalized:
                header_map[normalized] = col_idx

        return header_map

    def _row_to_dict(self, ws, row_idx: int, header_map: dict[str, int]) -> dict[str, Any]:
        """
        Превращает конкретную Excel-строку в dict по именам заголовков.
        """
        data: dict[str, Any] = {}
        for header, col_idx in header_map.items():
            data[header] = ws.cell(row=row_idx, column=col_idx).value
        return data

    def _is_empty_row(self, raw: dict[str, Any]) -> bool:
        """
        Полностью пустую строку игнорируем.
        """
        return all(v is None or str(v).strip() == "" for v in raw.values())

    # =========================================================
    # Парсинг и валидация строки
    # =========================================================

    def _parse_and_validate_row(self, raw: dict[str, Any]) -> ParsedRow:
        """
        Проверяет обязательные поля, числа, длины строк
        и возвращает ParsedRow в нормализованном виде.
        """
        # ---- обязательные текстовые поля ----
        product_sku = self._required_text(raw.get("product_sku"), "product_sku")
        product_name = self._required_text(raw.get("product_name"), "product_name")
        category_name = self._required_text(raw.get("category_name"), "category_name")
        seller_name = self._required_text(raw.get("seller_name"), "seller_name")
        variant_name = self._required_text(raw.get("variant_name"), "variant_name")

        # ---- обязательные денежные поля ----
        cost_price = self._decimal_value(raw.get("cost_price"), "cost_price")
        sale_price = self._decimal_value(raw.get("sale_price"), "sale_price")

        if cost_price < 0:
            raise ValueError("cost_price не может быть отрицательной")
        if sale_price < 0:
            raise ValueError("sale_price не может быть отрицательной")

        # ---- pack_size и stock ----
        pack_size = self._int_value(raw.get("pack_size"), default=1)
        if pack_size < 1:
            raise ValueError("pack_size должен быть >= 1")

        stock = self._int_value(raw.get("stock"), default=0)
        if stock < 0:
            raise ValueError("stock не может быть отрицательным")

        # ---- флаг активности ----
        is_active = self._bool_01_value(raw.get("is_active"), default=True)

        row = ParsedRow(
            product_sku=product_sku,
            product_name=product_name,
            product_description=self._optional_text(raw.get("product_description")),
            category_name=category_name,
            seller_name=seller_name,
            seller_city=self._optional_text(raw.get("seller_city")),
            unit=self._optional_text(raw.get("unit")) or "шт",
            main_image_url=self._optional_text(raw.get("main_image_url")),
            gallery_urls=self._split_urls(raw.get("gallery_urls")),
            youtube_url=self._optional_text(raw.get("youtube_url")),
            variant_name=variant_name,
            pack_size=pack_size,
            cost_price=cost_price,
            sale_price=sale_price,
            stock=stock,
            is_active=is_active,
        )

        # ---- проверки длины строк под ограничения БД ----
        if len(row.product_sku) > 64:
            raise ValueError("product_sku длиннее 64 символов")
        if len(row.product_name) > 255:
            raise ValueError("product_name длиннее 255 символов")
        if len(row.variant_name) > 120:
            raise ValueError("variant_name длиннее 120 символов")
        if len(row.unit) > 32:
            raise ValueError("unit длиннее 32 символов")
        if row.product_description and len(row.product_description) > 2000:
            raise ValueError("product_description длиннее 2000 символов")

        return row

    # =========================================================
    # Создание / поиск справочников
    # =========================================================

    def _get_or_create_category(self, category_name: str) -> tuple[Category, bool]:
        """
        Ищет категорию по имени без учета регистра.
        Если не найдено — создает новую категорию со slug.
        """
        cache_key = category_name.casefold()

        if cache_key in self._category_cache:
            return self._category_cache[cache_key], False

        category = (
            self.db.query(Category)
            .filter(func.lower(Category.name) == category_name.casefold())
            .first()
        )

        created = False
        if not category:
            slug = self._make_unique_category_slug(category_name)
            category = Category(name=category_name, slug=slug)
            self.db.add(category)
            self.db.flush()
            created = True

        self._category_cache[cache_key] = category
        return category, created

    def _get_or_create_seller(self, seller_name: str, seller_city: str | None) -> tuple[Seller, bool]:
        """
        Ищет продавца по имени без учета регистра.
        Если не найден — создает.
        Если найден, но city пустой, а в Excel city есть — дозаполняет.
        """
        cache_key = seller_name.casefold()

        if cache_key in self._seller_cache:
            seller = self._seller_cache[cache_key]
            if (not seller.city or not seller.city.strip()) and seller_city:
                seller.city = seller_city
                self.db.flush()
            return seller, False

        seller = (
            self.db.query(Seller)
            .filter(func.lower(Seller.name) == seller_name.casefold())
            .first()
        )

        created = False
        if not seller:
            seller = Seller(name=seller_name, city=(seller_city or "Не указан"))
            self.db.add(seller)
            self.db.flush()
            created = True
        elif (not seller.city or not seller.city.strip()) and seller_city:
            seller.city = seller_city
            self.db.flush()

        self._seller_cache[cache_key] = seller
        return seller, created

    # =========================================================
    # Создание / обновление Product и Variant
    # =========================================================

    def _get_or_create_product(
        self,
        row: ParsedRow,
        category_id: int,
        seller_id: int,
    ) -> tuple[Product, bool]:
        """
        Product определяется строго по SKU.

        Важный момент:
        новый Product сначала создается как объект,
        потом заполняются обязательные поля,
        и только после этого идет flush().
        Это защищает от NOT NULL ошибок в БД.
        """
        if row.product_sku in self._product_cache:
            product = self._product_cache[row.product_sku]
            created = False
        else:
            product = self.db.query(Product).filter(Product.sku == row.product_sku).first()
            created = product is None

            if created:
                product = Product(sku=row.product_sku)
                self.db.add(product)

            self._product_cache[row.product_sku] = product

        product.name = row.product_name
        product.description = row.product_description
        product.unit = row.unit or "шт"
        product.category_id = category_id
        product.seller_id = seller_id
        product.is_active = row.is_active

        self.db.flush()
        return product, created

    def _get_or_create_variant(self, product_id: int, row: ParsedRow) -> tuple[Variant, bool]:
        """
        Variant определяется по:
            product_id + variant_name + pack_size
        """
        variant = (
            self.db.query(Variant)
            .filter(
                Variant.product_id == product_id,
                func.lower(Variant.name) == row.variant_name.casefold(),
                Variant.pack_size == row.pack_size,
            )
            .first()
        )

        created = variant is None
        if created:
            variant = Variant(product_id=product_id)
            self.db.add(variant)

        variant.name = row.variant_name
        variant.pack_size = row.pack_size
        variant.unit_price_net_cost = row.cost_price
        variant.unit_price = row.sale_price
        variant.stock = row.stock
        variant.is_active = row.is_active

        self.db.flush()
        return variant, created

    # =========================================================
    # Медиа
    # =========================================================

    def _process_product_media_once(
        self,
        product: Product,
        row: ParsedRow,
        row_idx: int,
        result: ImportResult,
    ) -> None:
        """
        Медиа для товара обрабатываем только один раз на SKU за весь импорт.

        Почему это нужно:
        один и тот же товар может идти в нескольких строках Excel
        (разные варианты), но медиа у товара общее.
        """
        sku_key = row.product_sku.casefold()

        if sku_key in self._media_processed_skus:
            # Медиа для этого SKU уже обработали ранее
            result.media_skipped += self._count_media_items(row)
            return

        self._media_processed_skus.add(sku_key)

        # Главное фото
        if row.main_image_url:
            try:
                self._replace_main_image(product, row.main_image_url)
                result.media_processed += 1
            except Exception as exc:
                result.add_warning(row_idx, f"Главное фото не сохранено: {exc}")

        # Галерея
        if row.gallery_urls:
            try:
                self._replace_gallery(product, row.gallery_urls)
                result.media_processed += len(row.gallery_urls)
            except Exception as exc:
                result.add_warning(row_idx, f"Галерея не сохранена: {exc}")

        # Видео
        if row.youtube_url:
            try:
                self._replace_video(product, row.youtube_url)
                result.media_processed += 1
            except Exception as exc:
                result.add_warning(row_idx, f"YouTube видео не сохранено: {exc}")

        self.db.flush()

    def _replace_main_image(self, product: Product, image_url: str) -> None:
        """
        Полностью заменяет главное фото товара:
        - удаляет старый файл (если был)
        - скачивает новый
        - сохраняет имя файла в Product.image
        """
        content, ext = download_image_bytes(image_url)
        new_filename = save_main_image_bytes(content, ext)

        if product.image:
            delete_main_image_file(product.image)

        product.image = new_filename

    def _replace_gallery(self, product: Product, gallery_urls: list[str]) -> None:
        """
        Полностью заменяет галерею товара:
        - удаляет старые записи ProductImage
        - очищает папку gallery
        - скачивает и сохраняет новые файлы
        - создает новые ProductImage с sort_order
        """
        # 1. удалить старые записи из БД
        for img in list(product.images):
            self.db.delete(img)

        # 2. очистить папку gallery на диске
        clear_product_gallery_dir(product.id)

        self.db.flush()

        # 3. записать новые файлы
        for index, url in enumerate(gallery_urls):
            content, ext = download_image_bytes(url)
            filename = save_gallery_image_bytes(product.id, content, ext)

            self.db.add(
                ProductImage(
                    product_id=product.id,
                    image_url=filename,
                    sort_order=index,
                )
            )

    def _replace_video(self, product: Product, youtube_url: str) -> None:
        """
        Полностью заменяет видео товара:
        - удаляет старые ProductVideo
        - сохраняет новое видео
        В БД сохраняется именно YouTube ID, чтобы это было
        совместимо с твоей ручной админкой.
        """
        for video in list(product.videos):
            self.db.delete(video)

        self.db.flush()

        youtube_id = extract_youtube_id(youtube_url)
        if not youtube_id:
            raise ValueError("Не удалось извлечь YouTube ID")

        self.db.add(
            ProductVideo(
                product_id=product.id,
                video_url=youtube_id,
                title=None,
                sort_order=0,
            )
        )

    def _count_media_items(self, row: ParsedRow) -> int:
        """
        Сколько медиа-элементов было указано в строке.
        Используется для статистики media_skipped,
        когда SKU уже был обработан ранее в этом же файле.
        """
        total = 0
        if row.main_image_url:
            total += 1
        if row.gallery_urls:
            total += len(row.gallery_urls)
        if row.youtube_url:
            total += 1
        return total

    # =========================================================
    # Вспомогательные методы
    # =========================================================

    def _make_unique_category_slug(self, value: str) -> str:
        """
        Генерация уникального slug для новой категории.
        """
        base = self._slugify(value) or "category"
        slug = base
        index = 2

        while self.db.query(Category).filter(Category.slug == slug).first():
            slug = f"{base}-{index}"
            index += 1

        return slug

    @staticmethod
    def _normalize_header(value: str) -> str:
        """
        Нормализация заголовка Excel:
        'Product SKU' -> 'product_sku'
        """
        return "_".join(value.strip().lower().split())

    @staticmethod
    def _normalize_text(value: str) -> str:
        """
        Нормализация текста:
        удаляем лишние пробелы внутри и по краям.
        """
        return " ".join(value.strip().split())

    def _required_text(self, value: Any, field_name: str) -> str:
        """
        Получить обязательное текстовое поле.
        Если пусто — бросаем ValueError.
        """
        text = self._optional_text(value)
        if not text:
            raise ValueError(f"Поле {field_name} обязательно")
        return text

    def _optional_text(self, value: Any) -> str | None:
        """
        Получить необязательное текстовое поле.
        Если пусто — возвращаем None.
        """
        if value is None:
            return None
        text = self._normalize_text(str(value))
        return text or None

    @staticmethod
    def _decimal_value(value: Any, field_name: str) -> Decimal:
        """
        Парсинг Decimal:
        поддерживаем и '123.45', и '123,45'
        """
        if value is None or str(value).strip() == "":
            raise ValueError(f"Поле {field_name} обязательно")

        text = str(value).strip().replace(" ", "").replace(",", ".")
        try:
            return Decimal(text)
        except (InvalidOperation, ValueError):
            raise ValueError(f"Поле {field_name} должно быть числом")

    @staticmethod
    def _int_value(value: Any, default: int) -> int:
        """
        Парсинг целого числа.
        Пустое значение заменяем на default.
        """
        if value is None or str(value).strip() == "":
            return default
        try:
            return int(float(str(value).strip().replace(",", ".")))
        except (ValueError, TypeError):
            raise ValueError("Ожидалось целое число")

    @staticmethod
    def _bool_01_value(value: Any, default: bool) -> bool:
        """
        Парсинг булевого значения из:
        1 / 0 / 1.0 / 0.0 / true / false / yes / no / да / нет
        """
        if value is None or str(value).strip() == "":
            return default

        text = str(value).strip().lower()

    # 🔹 сначала пробуем как число (Excel чаще всего так дает)
        try:
            num = int(float(text))
            if num == 1:
                return True
            if num == 0:
                return False
        except (ValueError, TypeError):
            pass
    # 🔹 fallback на текстовые значения
        if text in {"true", "yes", "да"}:
            return True
        if text in {"false", "no", "нет"}:
            return False

        raise ValueError("is_active должен быть 0 или 1")

    @staticmethod
    def _split_urls(value: Any) -> list[str]:
        """
        Разделение строки URL по запятым.
        Пример:
            "url1, url2, url3" -> ["url1", "url2", "url3"]
        """
        if value is None:
            return []

        text = str(value).strip()
        if not text:
            return []

        return [part.strip() for part in text.split(",") if part.strip()]

    @staticmethod
    def _slugify(value: str) -> str:
        """
        Простая генерация slug:
        - буквы/цифры оставляем
        - остальные символы превращаем в "-"
        - подряд идущие "-" схлопываем
        """
        value = value.strip().lower()
        allowed = []
        prev_dash = False

        for ch in value:
            if ch.isalnum():
                allowed.append(ch)
                prev_dash = False
            else:
                if not prev_dash:
                    allowed.append("-")
                    prev_dash = True

        slug = "".join(allowed).strip("-")
        return slug